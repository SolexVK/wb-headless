#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Пересборка исходников в чистом виде: «GTIN коды» и «Инвентаризация».

    python3 rebuild_sources.py --nomenclature Т1.xlsx --gtin Т2.xlsx \
                               --inventory Т3.xlsx --out-dir .

Зачем это отдельно от build_kizy.py: там мы СОБИРАЕМ итог, а здесь наводим
порядок в самих исходниках — сводим артикулы к написанию из номенклатуры,
цвета к единому справочнику, размеры к одному ряду. Оба файла после этого
сходятся друг с другом ключом «артикул продавца», и нечёткое сопоставление
в следующий раз уже не понадобится.

GTIN коды      — лист на каждый номер модели из номенклатуры;
                 колонки: Артикул продавца · Цвет · Размер · GTIN.
Инвентаризация — лист на каждый номер, где инвентаризация проводилась;
                 колонки: Артикул продавца · Цвет · Статус КИЗ · размеры.

Суммы количеств не меняются — скрипт проверяет это сам и падает, если разошлись.
"""
import argparse, os, sys
from collections import defaultdict

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import kiz_lib as K
from colors import Dict_, tidy, from_article


def size_out(s):
    return str(s).strip().upper().replace('Х', 'X').replace('М', 'M')


def read_matrix(path):
    """Инвентаризация как есть: строки-цвета целиком, включая нулевые.

    load_inventory() отдаёт только непустые ячейки — для пересборки этого мало,
    строки с нулевым остатком тоже должны остаться в таблице.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    out = {}
    for ws in wb.worksheets:
        no = K.art_no(ws.title)
        header, block, rows, sizes = None, None, [], []
        for row in ws.iter_rows(values_only=True):
            cells = [c for c in row if c not in (None, '')]
            if not cells:
                continue
            if len(cells) == 1 and isinstance(cells[0], str) and 'киз' in cells[0].lower():
                block = 'Без кизов' if cells[0].lower().strip().startswith('без') else 'С кизами'
                continue
            if all(isinstance(c, str) for c in cells) and any(K.looks_like_size(c) for c in cells):
                header = row
                sizes += [size_out(c) for c in cells]
                continue
            if header is None or row[0] in (None, '') or not str(row[0]).strip():
                continue
            qty = {}
            for ci, v in enumerate(row):
                if ci == 0 or not isinstance(v, (int, float)):
                    continue
                s = header[ci] if ci < len(header) else None
                if s in (None, ''):
                    continue
                qty[size_out(s)] = qty.get(size_out(s), 0) + int(v)
            rows.append(dict(color=str(row[0]).strip(), block=block or 'Без кизов', qty=qty))
        out[no] = dict(rows=rows, sizes=sorted(set(sizes), key=K.size_sort_key))
    return out


def assign_gtin(t1, t2):
    """GTIN к строкам номенклатуры, по одному коду на позицию.

    Внутри группы «номер + размер» назначаем жадно от лучшего совпадения цвета:
    иначе один и тот же код мог бы достаться двум похожим цветам.
    """
    b1, b2 = defaultdict(list), defaultdict(list)
    for r in t1: b1[(r['no'], r['sz'])].append(r)
    for r in t2: b2[(r['no'], r['sz'])].append(r)
    gtin_of, used = {}, set()
    for key, rows in b1.items():
        cands = b2.get(key, [])
        pairs = sorted(((K.color_score(a['ct'], b['ct']), i, j)
                        for i, a in enumerate(rows) for j, b in enumerate(cands)),
                       key=lambda x: (-x[0][0], -x[0][1]))
        ta, tb = set(), set()
        for sc, i, j in pairs:
            if sc[0] < 0.5 or i in ta or j in tb:
                continue
            ta.add(i); tb.add(j)
            gtin_of[id(rows[i])] = cands[j]['gtin']
            used.add(cands[j]['gtin'])
    return gtin_of, [r for r in t2 if r['gtin'] not in used]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--nomenclature', required=True)
    ap.add_argument('--gtin', required=True)
    ap.add_argument('--inventory', required=True)
    ap.add_argument('--out-dir', default='.')
    ap.add_argument('--merge-kiz', action='store_true',
                    help='без колонки «Статус КИЗ» — блоки «с кизами/без кизов» складываются')
    ap.add_argument('--blank-zeros', action='store_true',
                    help='оставлять нулевые остатки пустыми, как в оригинале')
    a = ap.parse_args()

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    t1 = K.load_nomenclature(a.nomenclature)
    t2, _ = K.load_gtin(a.gtin)
    t3 = K.load_inventory(a.inventory)

    # ---- единый справочник цветов (инвентаризация > GTIN > артикул) ----
    D = Dict_()
    for r in t3: D.add(r['color'], 1)
    for r in t2:
        if r['color']: D.add(r['color'], 2)
    for art in {r['art'] for r in t1}: D.add(from_article(art), 3)
    D.build()

    gtin_of, orphans = assign_gtin(t1, t2)

    by_no = defaultdict(dict)
    for r in t1: by_no[r['no']].setdefault(r['art'], r['ct'])

    def art_for(no, color):
        tok = K.color_tokens(color)
        best, bs = None, (0, 0)
        for art, ct in by_no.get(no, {}).items():
            sc = K.color_score(tok, ct)
            if sc > bs: bs, best = sc, art
        return best if bs[0] >= 0.5 else None

    # ---- оформление ----
    HF, HFONT = PatternFill('solid', fgColor='C9DAF0'), Font(bold=True, color='16202A')
    A_H = Alignment(horizontal='center', vertical='center', wrap_text=True)
    A_T = Alignment(horizontal='left', vertical='center')
    A_N = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='BFCBD8')
    BD = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style(ws, ncols, text_cols, widths):
        for c in ws[1]:
            c.font, c.fill, c.alignment, c.border = HFONT, HF, A_H, BD
        for row in ws.iter_rows(min_row=2, max_col=ncols):
            for i, c in enumerate(row, 1):
                c.border = BD
                c.alignment = A_T if i in text_cols else A_N
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = 'A2'

    # ---- файл 1: GTIN коды ----
    wg = openpyxl.Workbook(); wg.remove(wg.active)
    for no in sorted(by_no):
        ws = wg.create_sheet(no)
        ws.append(['Артикул продавца', 'Цвет', 'Размер', 'GTIN'])
        for r in sorted((x for x in t1 if x['no'] == no),
                        key=lambda x: (x['art'], K.size_sort_key(x['size']))):
            ws.append([r['art'], D.get(from_article(r['art'])),
                       size_out(r['size']), gtin_of.get(id(r))])
        for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
            for c in row: c.number_format = '@'
        style(ws, 4, {1, 2, 3}, [32, 22, 11, 18])
    p_gtin = os.path.join(a.out_dir, 'GTIN коды.xlsx')
    wg.save(p_gtin)

    # ---- файл 2: Инвентаризация ----
    M = read_matrix(a.inventory)
    wi = openpyxl.Workbook(); wi.remove(wi.active)
    total = 0
    for no in sorted(M):
        sizes = sorted(set(M[no]['sizes']) | {size_out(r['size']) for r in t1 if r['no'] == no},
                       key=K.size_sort_key)
        head = ['Артикул продавца', 'Цвет'] + ([] if a.merge_kiz else ['Статус КИЗ']) + sizes
        ws = wi.create_sheet(no)
        ws.append(head)
        merged = {}
        for r in M[no]['rows']:
            art = art_for(no, r['color'])
            key = (art, r['color']) if a.merge_kiz else (art, r['color'], r['block'])
            acc = merged.setdefault(key, dict(block=r['block'], qty=defaultdict(int)))
            for s, q in r['qty'].items():
                acc['qty'][s] += q
        for key, acc in merged.items():
            line = [key[0], D.get(key[1])] + ([] if a.merge_kiz else [acc['block']])
            for s in sizes:
                q = acc['qty'].get(s, 0)
                total += q
                line.append(None if (a.blank_zeros and not q) else q)
            ws.append(line)
        ntext = 2 if a.merge_kiz else 3
        style(ws, len(head), set(range(1, ntext + 1)),
              [32, 22] + ([] if a.merge_kiz else [13]) + [7] * len(sizes))
    p_inv = os.path.join(a.out_dir, 'Инвентаризация.xlsx')
    wi.save(p_inv)

    src_total = sum(r['qty'] for r in t3)
    print(f'GTIN коды      : {len(by_no)} листов, {len(t1)} строк, кодов {len(gtin_of)}, '
          f'пусто {len(t1) - len(gtin_of)}  -> {p_gtin}')
    print(f'Инвентаризация : {len(M)} листов, {total} шт  -> {p_inv}')
    if orphans:
        print(f'\nGTIN без пары в номенклатуре: {len(orphans)}')
        for r in sorted(orphans, key=lambda x: (x['no'], str(x['color']), x['size'])):
            print(f'   {r["no"]}  {(r["color"] or r["art"]):14s} р.{r["size"]:6s} {r["gtin"]}')
    if total != src_total:
        print(f'\n!! СУММА НЕ СОШЛАСЬ: {total} против {src_total} в оригинале')
        return 1
    print(f'\nСумма сошлась с оригиналом: {total} шт')
    return 0


if __name__ == '__main__':
    sys.exit(main())
