#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Разведка исходников перед сборкой: что за файл, как он устроен, где сюрпризы.

    python3 inspect_inputs.py файл1.xlsx [файл2.xlsx ...]

Тип файла определяется по содержимому, а не по имени, потому что имена
приходят разные («GTIN коды 2», «Инвентаризация 30.08.26»). Запускай это
первым делом на каждом новом наборе: если структура уехала, ты увидишь это
здесь, а не в виде тихо пустых колонок в итоге.
"""
import os, sys
from collections import Counter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import kiz_lib as K


def sniff(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    txt = 0
    for ws in wb.worksheets[:3]:
        for row in ws.iter_rows(max_row=30, values_only=True):
            for c in row:
                if isinstance(c, str) and c.strip().lower().startswith('арт'):
                    txt += 1
    if txt >= 3:
        return 'gtin'
    hdr = [str(c).strip().lower() for c in next(wb.worksheets[0].iter_rows(values_only=True))
           if c is not None]
    # шаблон проверяем раньше номенклатуры: у него тоже есть «Артикул продавца»
    if any('gtin' in h for h in hdr) and any('количество' in h for h in hdr):
        return 'target'
    if any('артикул продавца' in h for h in hdr):
        return 'nomenclature'
    return 'inventory'


def main():
    for path in sys.argv[1:]:
        kind = sniff(path)
        print(f'\n{"="*70}\n{os.path.basename(path)}  ->  {kind}\n{"="*70}')
        try:
            if kind == 'nomenclature':
                r = K.load_nomenclature(path)
                print(f'строк {len(r)} · артикулов {len({x["art"] for x in r})} · '
                      f'баркодов уник. {len({x["bc"] for x in r})}')
                print('размеры:', dict(Counter(x['size'] for x in r).most_common()))
                bad = [x for x in r if not x['bc'] or not x['wb']]
                if bad:
                    print(f'!! без баркода/артикула WB: {len(bad)} строк')
                print('номера моделей:', sorted({x['no'] for x in r}))

            elif kind == 'gtin':
                r, bad = K.load_gtin(path)
                print(f'кодов {len(r)} · уникальных {len({x["gtin"] for x in r})} · '
                      f'листов {len({x["sheet"] for x in r})}')
                dup = [g for g, n in Counter(x['gtin'] for x in r).items() if n > 1]
                if dup:
                    print(f'!! ДУБЛИ GTIN: {dup[:10]}')
                if bad:
                    print(f'!! не разобрано строк: {len(bad)}')
                    for b in bad[:10]:
                        print('   ', b)
                print('размеры:', dict(Counter(x['size'] for x in r).most_common()))
                print('номера моделей:', sorted({x['no'] for x in r}))

            elif kind == 'inventory':
                r = K.load_inventory(path)
                print(f'ячеек {len(r)} · всего {sum(x["qty"] for x in r)} шт · '
                      f'листов {len({x["sheet"] for x in r})}')
                blocks = Counter()
                for x in r:
                    blocks[x['block'] or 'без кизов'] += x['qty']
                print('блоки, шт:', dict(blocks))
                print('размеры:', dict(Counter(x['size'] for x in r).most_common()))
                print('по листам:', {s: sum(x['qty'] for x in r if x['sheet'] == s)
                                     for s in sorted({x['sheet'] for x in r})})
                print('цвета:', sorted({x['color'] for x in r}))

            else:
                import openpyxl
                ws = openpyxl.load_workbook(path).worksheets[0]
                print('шапка:', [c.value for c in ws[1]])
        except Exception as e:
            print(f'!! не разобрался: {type(e).__name__}: {e}')


if __name__ == '__main__':
    main()
