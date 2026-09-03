#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Сборка итоговой таблицы «для кизов».

    python3 build_kizy.py --nomenclature Т1.xlsx --gtin Т2.xlsx \
                          --inventory Т3.xlsx --out для_кизов_ИТОГ.xlsx

Ведущая таблица — инвентаризация: одна строка результата на каждую непустую
ячейку остатка. Всё остальное (артикул продавца, артикул WB, баркод, GTIN)
подтягивается к ней по ключу «номер модели + цвет + размер».

Скрипт НИЧЕГО не выдумывает: если пары не нашлось, ячейка остаётся пустой,
а строка попадает в отчёт. Отчёт (--report) и раздел «ТРЕБУЕТ ВНИМАНИЯ»
в консоли — это то, что нужно показать пользователю.
"""
import argparse, json, os, sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import kiz_lib as K


def build(args):
    t1 = K.load_nomenclature(args.nomenclature)
    t2, t2_bad = K.load_gtin(args.gtin, width=args.gtin_width)
    t3 = K.load_inventory(args.inventory)

    ov = json.load(open(args.overrides, encoding='utf-8')) if args.overrides else {}
    color_map = ov.get('color_map', {})      # "014|пудра" -> "розовый"
    force_art = ov.get('force_art', {})      # "016|Радуга" -> "016 МС сердечко радуга"

    i1, i2 = K.index_by(t1), K.index_by(t2)
    by_art = {r['art']: r for r in t1}

    recs, notes = [], []
    for inv in t3:
        no, sz = inv['art_no'], K.norm_size(inv['size'])
        key = f"{no}|{inv['color'].strip().lower()}"
        color = color_map.get(key, inv['color'])
        tok = K.color_tokens(color)

        m1, s1, n1 = K.best_match(i1.get((no, sz), []), tok, args.threshold)
        if key in force_art:
            pinned = [r for r in i1.get((no, sz), []) if r['art'] == force_art[key]]
            m1, s1, n1 = (pinned[0], 1.0, 1) if pinned else (m1, s1, n1)
        m2, s2, n2 = K.best_match(i2.get((no, sz), []), tok, args.threshold)

        status = []
        if not m1:
            status.append('нет в номенклатуре')
        if not m2:
            status.append('нет GTIN')
        if m1 and s1 < 1.0:
            status.append(f'цвет подобран нестрого ({s1:.2f})')
            notes.append((inv, m1['art'], s1, n1))

        recs.append(dict(
            art=m1['art'] if m1 else None,
            wb=m1['wb'] if m1 else None,
            gtin=m2['gtin'] if m2 else None,
            size=m1['size'] if m1 else inv['size'],
            bc=m1['bc'] if m1 else None,
            qty=inv['qty'],
            color=inv['color'], block=inv['block'] or 'без кизов',
            sheet=inv['sheet'], status='; '.join(status) or 'ок'))

    if args.exclude_kiz:
        recs = [r for r in recs if r['block'] != 'с кизами']
    recs.sort(key=lambda r: (r['sheet'], str(r['art']), K.size_sort_key(r['size'])))
    return recs, notes, t1, t2, t3, t2_bad, by_art


def write_xlsx(recs, path, single_sheet=False):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HDR = Font(bold=True)
    FILL = PatternFill('solid', fgColor='DDEBF7')
    WARN = PatternFill('solid', fgColor='FFF2CC')
    KIZ = PatternFill('solid', fgColor='E2EFDA')

    wb = openpyxl.Workbook()
    o = wb.active
    o.title = 'Worksheet'
    o.append(['Артикул продавца', 'Артикул', 'GTIN', 'Размер', 'Баркод', 'Количество'])
    for c in o[1]:
        c.font, c.fill = HDR, FILL
        c.alignment = Alignment(horizontal='center')
    for r in recs:
        o.append([r['art'], r['wb'], r['gtin'], r['size'], r['bc'], r['qty']])
    for row in o.iter_rows(min_row=2):
        row[2].number_format = '@'      # GTIN и баркод — только текстом,
        row[4].number_format = '@'      # иначе слетает ведущий ноль
    for col, w in zip('ABCDEF', [28, 13, 18, 10, 16, 13]):
        o.column_dimensions[col].width = w
    o.freeze_panes = 'A2'
    o.auto_filter.ref = f'A1:F{o.max_row}'

    if not single_sheet:
        p = wb.create_sheet('Проверка')
        p.append(['Артикул продавца', 'Артикул', 'GTIN', 'Размер', 'Баркод', 'Количество',
                  'Цвет (инвентаризация)', 'Блок', 'Лист инв.', 'Статус'])
        for c in p[1]:
            c.font, c.fill = HDR, FILL
            c.alignment = Alignment(horizontal='center')
        for r in recs:
            p.append([r['art'], r['wb'], r['gtin'], r['size'], r['bc'], r['qty'],
                      r['color'], r['block'], r['sheet'], r['status']])
        for row in p.iter_rows(min_row=2):
            row[2].number_format = '@'
            row[4].number_format = '@'
            if row[9].value != 'ок':
                for c in row:
                    c.fill = WARN
            elif row[7].value == 'с кизами':
                for c in row:
                    c.fill = KIZ
        for col, w in zip('ABCDEFGHIJ', [28, 13, 18, 10, 16, 13, 24, 13, 10, 22]):
            p.column_dimensions[col].width = w
        p.freeze_panes = 'A2'
        p.auto_filter.ref = f'A1:J{p.max_row}'
    wb.save(path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--nomenclature', required=True, help='Т1 «Перечень номенклатур»')
    ap.add_argument('--gtin', required=True, help='Т2 «GTIN коды»')
    ap.add_argument('--inventory', required=True, help='Т3 «Инвентаризация»')
    ap.add_argument('--out', required=True, help='итоговый .xlsx')
    ap.add_argument('--report', help='путь для JSON-отчёта')
    ap.add_argument('--overrides', help='JSON с color_map / force_art')
    ap.add_argument('--gtin-width', type=int, default=14,
                    help='14 = GTIN-14 текстом с ведущим нулём (по умолчанию), '
                         '13 = без ведущих нулей, 0 = как в исходнике')
    ap.add_argument('--exclude-kiz', action='store_true',
                    help='выбросить строки блока «С кизами»')
    ap.add_argument('--single-sheet', action='store_true',
                    help='только лист загрузки, без листа «Проверка»')
    ap.add_argument('--threshold', type=float, default=0.5,
                    help='минимальная схожесть цвета (0..1)')
    a = ap.parse_args()

    recs, notes, t1, t2, t3, t2_bad, by_art = build(a)
    write_xlsx(recs, a.out, a.single_sheet)

    inv_sum = sum(r['qty'] for r in t3)
    out_sum = sum(r['qty'] for r in recs)
    no_gtin = [r for r in recs if not r['gtin']]
    no_art = [r for r in recs if not r['art']]

    print(f'Т1 номенклатура : {len(t1)} строк, {len({r["art"] for r in t1})} артикулов')
    print(f'Т2 GTIN         : {len(t2)} кодов, {len({r["gtin"] for r in t2})} уникальных'
          + (f', НЕРАСПОЗНАНО {len(t2_bad)}' if t2_bad else ''))
    print(f'Т3 инвентаризация: {len(t3)} ячеек, {inv_sum} шт')
    print(f'\nИТОГ: {len(recs)} строк, {out_sum} шт -> {a.out}')
    print(f'  заполнено полностью : {len(recs)-len(no_gtin)-len(no_art)}')
    print(f'  без GTIN            : {len(no_gtin)} ({sum(r["qty"] for r in no_gtin)} шт)')
    print(f'  нет в номенклатуре  : {len(no_art)} ({sum(r["qty"] for r in no_art)} шт)')

    if no_gtin or no_art or notes:
        print('\n=== ТРЕБУЕТ ВНИМАНИЯ (показать пользователю) ===')
    for r in no_art:
        print(f'  [нет в номенклатуре] лист {r["sheet"]} · {r["color"]} · р.{r["size"]} = {r["qty"]}')
    for r in no_gtin:
        print(f'  [нет GTIN] {r["art"] or r["color"]} · р.{r["size"]} = {r["qty"]}')
    for inv, art, s, n in notes:
        print(f'  [подтвердить цвет] лист {inv["sheet"]} «{inv["color"]}» -> «{art}» '
              f'(схожесть {s:.2f}, кандидатов {n})')

    if a.report:
        json.dump(dict(
            totals=dict(rows=len(recs), qty=out_sum, inventory_qty=inv_sum,
                        filled=len(recs) - len(no_gtin) - len(no_art)),
            missing_gtin=[dict(art=r['art'], size=r['size'], qty=r['qty'],
                               color=r['color'], sheet=r['sheet']) for r in no_gtin],
            missing_nomenclature=[dict(color=r['color'], size=r['size'], qty=r['qty'],
                                       sheet=r['sheet']) for r in no_art],
            fuzzy_colors=[dict(sheet=i['sheet'], color=i['color'], matched=art,
                               score=round(s, 2), candidates=n) for i, art, s, n in notes],
            unparsed_gtin_rows=[dict(sheet=b[0], row=b[1]) for b in t2_bad],
        ), open(a.report, 'w', encoding='utf-8'), ensure_ascii=False, indent=2)
        print(f'\nОтчёт: {a.report}')

    if out_sum != inv_sum and not a.exclude_kiz:
        print(f'\n!! ВНИМАНИЕ: сумма в итоге ({out_sum}) != сумме в инвентаризации ({inv_sum})')
        return 1
    return 0


if __name__ == '__main__':
    sys.exit(main())
