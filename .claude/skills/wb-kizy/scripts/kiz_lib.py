# -*- coding: utf-8 -*-
"""
Библиотека для сборки таблицы «для кизов» из трёх выгрузок WB.

Все три исходника заполняются людьми вручную, поэтому ни на позицию колонки,
ни на точное написание цвета/размера полагаться нельзя. Отсюда два принципа:

1. Структуру определяем ПО СОДЕРЖИМОМУ ячеек, а не по номеру колонки.
2. Цвета и размеры сравниваем нечётко (токены + префиксы), а не строкой в строку.

Связующий ключ между таблицами — НЕ строка артикула (она в каждой таблице
пишется по-своему), а тройка: 3-значный номер артикула + цвет + размер.
"""
import re
from collections import defaultdict

# ----------------------------------------------------------------- размеры ---

_CYR2LAT = str.maketrans({'Х': 'X', 'М': 'M', 'С': 'C', 'А': 'A', 'Е': 'E', 'О': 'O'})


def norm_size(s):
    """'XXL'/'2XL' -> '2XL'; 'S 40 - 42' -> 'S'; кириллическая 'М' -> 'M'.

    Приводим к одному написанию только для СРАВНЕНИЯ. В итоговую таблицу
    размер выводим в том виде, в каком он записан в перечне номенклатур,
    потому что именно он считается эталонным для загрузки.
    """
    if s is None:
        return None
    s = str(s).strip().upper().translate(_CYR2LAT).strip()
    if re.fullmatch(r'\d+\s*-\s*\d+', s):          # 42-50, 52-58
        return re.sub(r'\s+', '', s)
    m = re.match(r'^([XSMLA0-9]+?)\s+\d+\s*-\s*\d+$', s)   # 'S 40 - 42' -> 'S'
    if m:
        s = m.group(1)
    s = s.replace(' ', '')
    if s == 'XXL':
        s = '2XL'
    if s == 'XXXL':
        s = '3XL'
    return s


SIZE_ORDER = ['XS', 'S', 'M', 'L', 'XL', 'XXL', '2XL', '3XL', '4XL',
              '5XL', '6XL', '7XL', '42-50', '52-58']


def size_sort_key(s):
    s = str(s)
    n = norm_size(s)
    for cand in (s, n):
        if cand in SIZE_ORDER:
            return (SIZE_ORDER.index(cand), s)
    return (99, s)


def looks_like_size(v):
    """Похоже ли значение на обозначение размера (для поиска строки-шапки)."""
    if v is None:
        return False
    s = norm_size(v)
    return bool(re.fullmatch(r'XS|S|M|L|XL|[2-9]XL|\d+-\d+', s or ''))


# ------------------------------------------------------------------- цвета ---

# Служебные куски артикулов (коды моделей, суффиксы партий, слова-описания).
# Если их не выбросить, «014 рубашка муслин белый» и «Белый» не сойдутся.
STOP = {
    'ип', 'ов', 'вс', 'пк', 'мк', 'лк', 'рм', 'мр', 'мс', 'рмп', 'ржп', 'ржк',
    'ржм', 'рмж', 'ржвп', 'рвп', 'рж', 'рмо', 'мпт', 'мпр', 'нср', 'нрс', 'пшр',
    'рмк', 'ржко', 'рмм', 'св', 'к', 'твид', 'рубашка', 'муслин', 'см', 'мм',
    'с', 'м', 'арт', 'цвет', 'р',
}

_END = re.compile(r'(ый|ий|ой|ая|яя|ое|ее|ые|ие|о|а|я|ь|е|й)+$')


def _stem(w):
    """Грубая основа слова: 'белая'/'белый'/'БЕЛЫЙ' -> 'бел'.

    Короткие слова не режем — у них окончание неотличимо от корня
    ('кофе', 'хаки', 'лен').
    """
    return _END.sub('', w) if len(w) > 4 else w


def color_tokens(c):
    """Строка цвета -> отсортированный набор основ значимых слов."""
    if c is None:
        return []
    s = str(c).lower().replace('ё', 'е')
    s = re.sub(r'[\-_/\\,.«»"\'()\[\]]+', ' ', s)   # кавычки важны: белый "Тюльпан"
    s = re.sub(r'\d+', '', s)
    out = []
    for w in s.split():
        if not w or w in STOP:
            continue
        t = _stem(w)
        if len(t) >= 3:
            out.append(t)
    return sorted(set(out))


def _tok_match(a, b):
    n = min(len(a), len(b))
    return a == b if n < 3 else a[:n] == b[:n]


def color_score(a, b):
    """(основной, дополнительный) — насколько похожи два набора токенов.

    Основной = доля совпавших от более длинного набора, дополнительный —
    от более короткого. Второй нужен как тай-брейк: «Радуга сердечки» одинаково
    (0.5) похожа на «радуга» и на «красный сердечко», но с «радуга» совпадает
    ВЕСЬ короткий набор (1.0), а с «красный сердечко» — половина.
    """
    if not a or not b:
        return (0.0, 0.0)
    used, hit = set(), 0
    for t in a:
        for j, u in enumerate(b):
            if j not in used and _tok_match(t, u):
                used.add(j)
                hit += 1
                break
    return (hit / max(len(a), len(b)), hit / min(len(a), len(b)))


def art_no(a):
    """3-значный номер модели из любой записи артикула."""
    m = re.search(r'(\d{3})', str(a))
    return m.group(1) if m else None


def art_color_tokens(a):
    """Цветовые токены из строки артикула: всё после 3-значного номера."""
    return color_tokens(re.sub(r'^\D*\d{3}', '', str(a)))


# ------------------------------------------------------------------ чтение ---

def _rows(ws):
    return [list(r) for r in ws.iter_rows(values_only=True)]


def _find_col(header, *names, exclude=()):
    """Индекс колонки по названию (регистронезависимо).

    Сначала пробуем точное совпадение и только потом подстроку: иначе
    «Размер» находится в «Код размера (chrt_id)», и в итог уезжает chrt_id.
    exclude отсеивает заголовки-ловушки на этапе поиска по подстроке.
    """
    norm = ['' if h is None else str(h).strip().lower() for h in header]
    for n in names:
        for i, h in enumerate(norm):
            if h == n:
                return i
    for n in names:
        for i, h in enumerate(norm):
            if n in h and not any(x in h for x in exclude):
                return i
    return None


def load_nomenclature(path):
    """Т1 «Перечень номенклатур» — один лист, шапка в первой непустой строке.

    Колонки ищем по названию: файл может приехать с другим порядком или с
    лишними колонками (Бренд, Предмет, Объём, Состав).
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    rows = _rows(ws)
    hi = next(i for i, r in enumerate(rows) if any(c not in (None, '') for c in r))
    header = rows[hi]
    ci = {
        'art':  _find_col(header, 'артикул продавца', 'артикул пост'),
        'wb':   _find_col(header, 'артикул wb', 'артикул вб', 'номенклатур',
                          exclude=('продавца', 'постав')),
        'size': _find_col(header, 'размер', exclude=('код', 'chrt')),
        'bc':   _find_col(header, 'баркод', 'штрихкод'),
    }
    if ci['art'] is None or ci['size'] is None:
        raise SystemExit(f'{path}: не нашёл колонки «Артикул продавца» / «Размер». '
                         f'Шапка: {header}')
    if ci['wb'] is None:   # «Артикул WB» мог называться просто «Артикул»
        ci['wb'] = _find_col(header, 'артикул', exclude=('продавца', 'постав'))
    out = []
    for r in rows[hi + 1:]:
        if not any(c not in (None, '') for c in r):
            continue
        art = r[ci['art']]
        if art in (None, ''):
            continue
        size = r[ci['size']]
        bc = r[ci['bc']] if ci['bc'] is not None else None
        out.append(dict(
            art=str(art).strip(),
            wb=r[ci['wb']] if ci['wb'] is not None else None,
            size=str(size).strip() if size is not None else None,
            bc=str(bc).strip() if bc is not None else None,
            no=art_no(art), sz=norm_size(size), ct=art_color_tokens(art)))
    return out


_RE_FULL = re.compile(r'^Арт\.?\s*(.*?)\s*,\s*цвет\s*:\s*(.*?)\s*,\s*р\.?\s*(.*)$', re.I)
_RE_SHORT = re.compile(r'^Арт\.?\s*(.*?)\s*,\s*р\.?\s*(.*)$', re.I)


def norm_gtin(v, width=14):
    """Код к единому виду. width=14 — GTIN-14 с ведущим нулём (нужен Честному
    Знаку), width=13 — без ведущих нулей, width=0 — как в исходнике."""
    s = str(v).strip()
    if s.endswith('.0'):
        s = s[:-2]
    if not width:
        return s
    if width < 14:
        s = s.lstrip('0') or '0'
    return s.zfill(width)


def load_gtin(path, width=14):
    """Т2 «GTIN коды» — много листов, на каждом описание + код.

    Колонки местами переставлены, а на одном листе код может лежать в разных
    колонках у разных строк, поэтому в каждой строке ищем ячейку с текстом
    «Арт...» и ячейку, которая целиком состоит из 8–14 цифр.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    out, bad = [], []
    for ws in wb.worksheets:
        for ri, row in enumerate(_rows(ws), 1):
            cells = [c for c in row if c not in (None, '')]
            if not cells:
                continue
            desc = gtin = None
            for c in cells:
                s = str(c).strip()
                if s.lower().startswith('арт'):
                    desc = s
                elif re.fullmatch(r'\d{8,14}(\.0)?', s):
                    gtin = s
            if desc is None or gtin is None:
                bad.append((ws.title, ri, cells))
                continue
            m = _RE_FULL.match(desc)
            if m:
                art, color, size = m.group(1), m.group(2), m.group(3)
                ct = color_tokens(color)
            else:
                m = _RE_SHORT.match(desc)
                if not m:
                    bad.append((ws.title, ri, desc))
                    continue
                art, color, size = m.group(1), None, m.group(2)
                ct = art_color_tokens(art)      # цвет зашит в артикул
            out.append(dict(sheet=ws.title, row=ri, art=art.strip(),
                            color=(color.strip() if color else None),
                            size=size.strip(), gtin=norm_gtin(gtin, width),
                            no=art_no(art), sz=norm_size(size), ct=ct))
    return out, bad


def load_inventory(path):
    """Т3 «Инвентаризация» — на каждом листе матрица цвет × размер.

    Шапка не обязательно в первой строке, порядок размеров произвольный
    (местами L идёт перед M), а внутри листа могут быть подблоки
    «С кизами» / «Без кизов» со своими шапками. Поэтому идём построчно и
    держим текущую шапку и текущий блок как состояние.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    out = []
    for ws in wb.worksheets:
        no = art_no(ws.title)
        header, block = None, None
        for ri, row in enumerate(_rows(ws), 1):
            cells = [c for c in row if c not in (None, '')]
            if not cells:
                continue
            if len(cells) == 1 and isinstance(cells[0], str) and 'киз' in cells[0].lower():
                low = cells[0].lower().strip()
                block = 'без кизов' if low.startswith('без') else 'с кизами'
                continue
            if all(isinstance(c, str) for c in cells) and any(looks_like_size(c) for c in cells):
                header = row
                continue
            if header is None:
                continue
            color = row[0]
            if color is None or not str(color).strip():
                continue
            for ci, val in enumerate(row):
                if ci == 0 or not isinstance(val, (int, float)) or val in (None, ''):
                    continue
                size = header[ci] if ci < len(header) else None
                if size in (None, ''):
                    continue
                out.append(dict(sheet=ws.title, row=ri, art_no=no,
                                color=str(color).strip(), size=str(size).strip(),
                                qty=int(val), block=block))
    return out


# --------------------------------------------------------------- сведение ---

def index_by(records, no_key='no', sz_key='sz'):
    idx = defaultdict(list)
    for r in records:
        idx[(r[no_key], r[sz_key])].append(r)
    return idx


def best_match(cands, tokens, threshold=0.5):
    """Лучший кандидат по цвету. Возвращает (запись, score, сколько равных)."""
    if not cands or not tokens:
        return None, 0.0, 0
    scored = sorted(((color_score(tokens, c['ct']), c) for c in cands),
                    key=lambda x: (-x[0][0], -x[0][1]))
    top = scored[0][0]
    if top[0] < threshold:
        return None, 0.0, 0
    n = sum(1 for s, _ in scored if s == top)
    return scored[0][1], top[0] * top[1], n
