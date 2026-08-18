# scripts/fbs-geo-xlsx.py — Excel по «Географии продаж и возвратов» FBS.
# Вход:  <REPORTS_OUTPUT_DIR>/fbs-geo-service.json (снимок сервиса)
# Выход: <REPORTS_OUTPUT_DIR>/fbs-geo.xlsx
# Листы: «Вся РФ — регионы», «Вся РФ — округа», «Моск. FF — регионы».
import os
import json
from openpyxl import Workbook
from xlsx_kit import autofit, style_header, autofilter

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_DIR = os.path.abspath(os.environ['REPORTS_OUTPUT_DIR']) if os.environ.get('REPORTS_OUTPUT_DIR') else os.path.join(REPO, 'reports-output')
SNAP = os.path.join(OUT_DIR, 'fbs-geo-service.json')
OUT = os.path.join(OUT_DIR, 'fbs-geo.xlsx')

with open(SNAP, encoding='utf-8') as f:
    s = json.load(f)

GREEN = '1B965A'  # цвет заливки шапки листов географии


def sheet(ws, rows, with_okrug):
    head = (['Округ'] if with_okrug else []) + ['Регион' if with_okrug else 'Округ', 'Продано', 'Возвращено', '% возврата', 'Сумма ₽']
    ws.append(head)
    style_header(ws, fill=GREEN, wrap=True)
    for r in rows:
        name = r.get('region') or r.get('okrug') or '—'
        row = ([r.get('okrug', '')] if with_okrug else []) + [name, r.get('salesCount', 0), r.get('returnCount', 0), r.get('returnPct', 0) / 100.0, round(r.get('salesRub', 0))]
        ws.append(row)
        rr = ws.max_row
        ws.cell(rr, ws.max_column - 1).number_format = '0.0%'
    ws.freeze_panes = 'A2'
    autofilter(ws)
    autofit(ws)


def gsheet(ws, head, rows, cells):
    ws.append(head)
    style_header(ws, fill=GREEN, wrap=True)
    for r in rows:
        ws.append(cells(r))
        rr = ws.max_row
        for j, h in enumerate(head, 1):
            if '%' in str(h):
                # Как на листах регионов: храним ДОЛЮ (0.333), формат '0.0%' сам умножит
                # на 100. Раньше тут лежало сырое 33.3 с литеральным «%» — визуально то же,
                # но значение в ячейке было в 100× другого масштаба (ломало сортировку/формулы).
                cell = ws.cell(rr, j)
                cell.value = (cell.value or 0) / 100.0
                cell.number_format = '0.0%'
    ws.freeze_panes = 'A2'
    autofilter(ws)
    autofit(ws)


wb = Workbook()
ws = wb.active
ws.title = 'Вся РФ — регионы'
sheet(ws, s.get('scopes', {}).get('all', {}).get('byRegion', []), True)
sheet(wb.create_sheet('Вся РФ — округа'), s.get('scopes', {}).get('all', {}).get('byOkrug', []), False)
sheet(wb.create_sheet('Моск. FF — регионы'), s.get('scopes', {}).get('moscow', {}).get('byRegion', []), True)

# FBS: возвраты по ФФ отгрузки.
F = s.get('fbs', {}) or {}
gsheet(wb.create_sheet('FBS по ФФ'), ['ФФ отгрузки', 'Отгружено', 'Выкуплено', 'Возвращено', '% возв.', 'Сумма ₽'],
       F.get('byFF', []), lambda r: [r.get('ff', ''), r.get('shipped', 0), r.get('salesCount', 0), r.get('returnCount', 0), r.get('returnPct', 0), r.get('salesRub', 0)])
gsheet(wb.create_sheet('FBS ФФ×регион'), ['ФФ отгрузки', 'Округ', 'Регион', 'Выкуплено', 'Возвращено', '% возв.'],
       F.get('ffByRegion', []), lambda r: [r.get('ff', ''), r.get('okrug', ''), r.get('region', ''), r.get('salesCount', 0), r.get('returnCount', 0), r.get('returnPct', 0)])
gsheet(wb.create_sheet('FBS по дням'), ['Дата', 'Выкуплено', 'Возвращено'],
       F.get('byDay', []), lambda r: [r.get('date', ''), r.get('salesCount', 0), r.get('returnCount', 0)])
gsheet(wb.create_sheet('FBS по товарам'), ['Артикул', 'Осн. ФФ', 'Выкуплено', 'Возвращено', '% возв.', 'Сумма ₽'],
       F.get('byArticle', []), lambda r: [r.get('article', ''), r.get('ff', ''), r.get('salesCount', 0), r.get('returnCount', 0), r.get('returnPct', 0), r.get('salesRub', 0)])

wb.save(OUT)
print('OK', OUT)
