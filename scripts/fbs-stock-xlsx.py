# scripts/fbs-stock-xlsx.py — Excel по остаткам FBS (для веб-сервиса).
# Вход:  <REPORTS_OUTPUT_DIR>/fbs-stock-service.json (агрегированный снимок сервиса)
# Выход: <REPORTS_OUTPUT_DIR>/fbs-stock.xlsx
# Значения, не формулы (чтобы не зависеть от пересчёта в LibreOffice/Excel).
import os
import json
from openpyxl import Workbook
from openpyxl.styles import Font
from xlsx_kit import autofit, style_header

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_DIR = os.path.abspath(os.environ['REPORTS_OUTPUT_DIR']) if os.environ.get('REPORTS_OUTPUT_DIR') else os.path.join(REPO, 'reports-output')
SNAP = os.path.join(OUT_DIR, 'fbs-stock-service.json')
OUT = os.path.join(OUT_DIR, 'fbs-stock.xlsx')

with open(SNAP, encoding='utf-8') as f:
    s = json.load(f)

t = s.get('totals', {})

wb = Workbook()

# Лист «Сводка»
ws = wb.active
ws.title = 'Сводка'
ws.append(['Остатки FBS — снимок'])
ws['A1'].font = Font(bold=True, size=13)
ws.append(['Собрано (UTC)', s.get('generatedAt', '')])
ws.append(['Всего, шт', t.get('grandTotal', 0)])
ws.append(['Активных складов', t.get('activeWarehouses', 0)])
ws.append(['Артикул+цвет', t.get('articleCount', 0)])
ws.append([])
ws.append(['Фулфилмент', 'Остаток, шт', 'Позиций (SKU)'])
for c in ws[7]:
    c.font = Font(bold=True)
for w in s.get('warehouses', []):
    ws.append([w.get('name'), w.get('totalQuantity', 0), w.get('skuInStock', 0)])
autofit(ws, cap=40)

# Лист «Артикул+цвет × склад»
ws2 = wb.create_sheet('Артикул+цвет × склад')
cols = s.get('warehouseList', [])
ws2.append(['Артикул', 'Цвет'] + list(cols) + ['Итого'])
style_header(ws2)
for a in s.get('articles', []):
    bw = a.get('byWarehouse', {})
    ws2.append([a.get('articleNum'), a.get('variant')] + [bw.get(c, 0) for c in cols] + [a.get('total', 0)])
autofit(ws2, cap=40)
ws2.freeze_panes = 'A2'

wb.save(OUT)
print('OK', OUT)
