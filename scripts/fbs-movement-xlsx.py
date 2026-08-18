# scripts/fbs-movement-xlsx.py — Excel по «Движению заказов» FBS (для веб-сервиса).
# Вход:  <REPORTS_OUTPUT_DIR>/fbs-movement-service.json (снимок сервиса)
# Выход: <REPORTS_OUTPUT_DIR>/fbs-movement.xlsx
# Листы (за период отображения, последние N дней): Принято/Передано × шт/₽.
# День × фулфилмент, значения (не формулы), со столбцом «Итого».
import os
import json
from openpyxl import Workbook
from xlsx_kit import autofit, style_header

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_DIR = os.path.abspath(os.environ['REPORTS_OUTPUT_DIR']) if os.environ.get('REPORTS_OUTPUT_DIR') else os.path.join(REPO, 'reports-output')
SNAP = os.path.join(OUT_DIR, 'fbs-movement-service.json')
OUT = os.path.join(OUT_DIR, 'fbs-movement.xlsx')

with open(SNAP, encoding='utf-8') as f:
    s = json.load(f)

fulfillments = s.get('fulfillments', [])
N = int(s.get('days', 14))
series = s.get('series', [])[-N:]  # период отображения
COST = float(os.environ.get('COST_PER_UNIT', 620) or 620)


# basis: 'count' | 'money' (цена заказа) | 'moneyAvg' (ср. цена продажи) | 'cost' (count×COST)
def cellval(bf, name, basis):
    o = bf.get(name, {}) or {}
    if basis == 'cost':
        return round((o.get('count', 0) or 0) * COST, 2)
    return o.get(basis, 0) or 0


def sheet(ws, kind, basis):
    ws.append(['Дата'] + list(fulfillments) + ['Итого'])
    style_header(ws)
    for d in series:
        bf = (d.get(kind, {}) or {}).get('byFulfillment', {}) or {}
        row = [d.get('date')]
        total = 0
        for name in fulfillments:
            v = cellval(bf, name, basis)
            row.append(v)
            total += v
        row.append(total if basis == 'count' else round(total, 2))
        ws.append(row)
    autofit(ws, cap=40)
    ws.freeze_panes = 'B2'


wb = Workbook()
ws = wb.active
ws.title = 'Принято (шт)'
sheet(ws, 'accepted', 'count')
sheet(wb.create_sheet('Передано (шт)'), 'delivered', 'count')
sheet(wb.create_sheet('Передано себест (руб)'), 'delivered', 'cost')
sheet(wb.create_sheet('Передано продажа (руб)'), 'delivered', 'moneyAvg')
sheet(wb.create_sheet('Передано заказ (руб)'), 'delivered', 'money')

wb.save(OUT)
print('OK', OUT)
