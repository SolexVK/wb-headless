#!/usr/bin/env python3
# scripts/fbs-xlsx.py — Excel-отчёт подсорта FBS (size-level) из снимка расчёта.
#
# Вход:  reports-output/fbs-replenishment.json (npm run fbs:replenish)
# Выход: reports-output/fbs-podsort.xlsx
#
# Листы:
#   «Сводка»        — параметры расчёта + итоги (KPI).
#   «Сводная»       — подсорт по (артикул № × цвет × размер) × склады (+ Итого).
#   «Подсорт»       — детально склад×артикул×цвет×размер (остаток, спрос/дн, дни до 0,
#                     подсорт, статус). Автофильтр + подсветка статусов.
#   «Пробный завоз» — по (артикул × цвет × размер) × склады: seed=seedMin на размер;
#                     «✓» — на складе уже есть; тип новинка/докладка.
#
# Значения — результат согласованного движка (scripts/fbs-replenishment.mjs).
# Формулы не используются: файл открывается с готовыми числами без пересчёта.

import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
# REPORTS_OUTPUT_DIR — папка ввода/вывода (мультитенант-сервис: свой каталог на кабинет).
OUT_DIR = os.path.abspath(os.environ['REPORTS_OUTPUT_DIR']) if os.environ.get('REPORTS_OUTPUT_DIR') else os.path.join(REPO, 'reports-output')
SNAP = os.path.join(OUT_DIR, 'fbs-replenishment.json')
OUT = os.path.join(OUT_DIR, 'fbs-podsort.xlsx')

with open(SNAP, encoding='utf-8') as f:
    snap = json.load(f)
P = snap['params']
T = snap['totals']
whList = snap['warehouseList']

FONT = 'Arial'
INK = '12211A'
ACCENT = '127045'
HEAD_FILL = PatternFill('solid', fgColor='1B965A')
PARAM_FILL = PatternFill('solid', fgColor='FFF3C4')
thin = Side(style='thin', color='DDE7E0')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def hcell(c, text):
    c.value = text
    c.font = Font(name=FONT, bold=True, color='FFFFFF', size=10)
    c.fill = HEAD_FILL
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = BORDER


def base(c, val, bold=False, align='left', numfmt=None, color=INK):
    c.value = val
    c.font = Font(name=FONT, bold=bold, color=color, size=10)
    c.alignment = Alignment(horizontal=align, vertical='center')
    c.border = BORDER
    if numfmt:
        c.number_format = numfmt


def add_status_cf(sheet, col_letter, first, last):
    red = PatternFill('solid', fgColor='FBE7EA'); redF = Font(name=FONT, color='C43A50', bold=True)
    amb = PatternFill('solid', fgColor='FAF0DA'); ambF = Font(name=FONT, color='B7791F', bold=True)
    grn = PatternFill('solid', fgColor='E4F3EC'); grnF = Font(name=FONT, color='16875A')
    rng = f'{col_letter}{first}:{col_letter}{last}'
    for val, fill, fnt in [('нет остатка', red, redF), ('разрыв до поставки', red, redF),
                           ('риск разрыва', amb, ambF), ('ок', grn, grnF)]:
        sheet.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=[f'"{val}"'], fill=fill, font=fnt))


wb = Workbook()

# ── Лист «Сводка» ───────────────────────────────────────────────────────────
s = wb.active
s.title = 'Сводка'
s['A1'] = 'FBS · Подсорт по фулфилмент-складам (по размерам)'
s['A1'].font = Font(name=FONT, bold=True, size=15, color=ACCENT)
s['A2'] = 'Расчёт по каждому складу и размеру (только FF/FBS). Значения — результат согласованного расчёта.'
s['A2'].font = Font(name=FONT, size=9, color='54695D')
s['A3'] = 'Снимок: ' + snap.get('generatedAt', '')[:16].replace('T', ' ') + ' UTC'
s['A3'].font = Font(name=FONT, size=9, color='88998F')

params = [
    ('Окно скорости, дн', P['velocityDays']),
    ('Лид-тайм, дн (мин–макс)', f"{P['leadMin']}–{P['leadMax']}"),
    ('Запас (cover), дн', P['coverDays']),
    ('Горизонт заказа, дн', P['horizonDays']),
    ('Seed, шт на размер', P['seedMin']),
    ('История «был на FF», дн', P['historyDays']),
    ('Артикулы в работе', ', '.join(P.get('articles', [])) or 'все'),
]
r0 = 5
base(s.cell(r0, 1), 'Параметры', bold=True)
s.cell(r0, 1).font = Font(name=FONT, bold=True, size=11, color=INK)
for i, (lab, val) in enumerate(params):
    rr = r0 + 1 + i
    base(s.cell(rr, 1), lab, bold=True)
    c = s.cell(rr, 2); base(c, val); c.fill = PARAM_FILL

k0 = r0 + len(params) + 2
base(s.cell(k0, 1), 'Итоги', bold=True)
s.cell(k0, 1).font = Font(name=FONT, bold=True, size=11, color=INK)
pos_podsort = sum(len(w['rows']) for w in snap['warehouses'])
kpis = [
    ('Подсорт всего, шт', T['reorderUnits']),
    ('Строк в риске', T['riskRows']),
    ('Строк подсорта (склад×размер)', pos_podsort),
    ('Строк сводной (арт×цвет×размер)', T.get('pivotRows', len(snap.get('pivot', [])))),
    ('Пробный завоз, шт', T['seedUnits']),
    ('Строк завоза (новинок+докладок)', f"{T['seedRows']} ({T['seedNovelty']}+{T['seedRefill']})"),
    ('Складов (все зарегистр.)', T['registeredWarehouses']),
]
for i, (lab, val) in enumerate(kpis):
    rr = k0 + 1 + i
    base(s.cell(rr, 1), lab, bold=True)
    isnum = isinstance(val, (int, float))
    base(s.cell(rr, 2), val, bold=True, align='right' if isnum else 'left',
         numfmt='#,##0' if isnum else None, color=ACCENT)
s.column_dimensions['A'].width = 34
s.column_dimensions['B'].width = 26

# ── Лист «Сводная» (артикул № × цвет × размер) × склады ──────────────────────
sv = wb.create_sheet('Сводная')
head = ['Арт', 'Цвет', 'Размер'] + whList + ['Итого']
for j, h in enumerate(head, 1):
    hcell(sv.cell(1, j), h)
row = 2
first_wh = 4  # D
for r in snap.get('pivot', []):
    base(sv.cell(row, 1), r['articleNum'], align='center')
    base(sv.cell(row, 2), r['variant'])
    base(sv.cell(row, 3), r['techSize'], align='center')
    for wi, wn in enumerate(whList):
        v = r['byWarehouse'].get(wn, 0)
        base(sv.cell(row, first_wh + wi), (v if v else None), align='right', numfmt='#,##0')
    base(sv.cell(row, first_wh + len(whList)), r['total'], align='right', bold=True, numfmt='#,##0')
    row += 1
sv_last = row - 1
sv.freeze_panes = 'D2'
sv.auto_filter.ref = f'A1:{get_column_letter(first_wh + len(whList))}{sv_last}'
for col, wdt in [('A', 6), ('B', 22), ('C', 9)]:
    sv.column_dimensions[col].width = wdt
for wi in range(len(whList)):
    sv.column_dimensions[get_column_letter(first_wh + wi)].width = 13
sv.column_dimensions[get_column_letter(first_wh + len(whList))].width = 9

# ── Лист «Подсорт» (детально) ───────────────────────────────────────────────
p = wb.create_sheet('Подсорт')
cols = ['Склад', 'Арт', 'Цвет', 'Размер', 'nmID', 'Остаток', 'Спрос/дн', 'Дни до 0', 'Подсорт', 'Статус']
for j, h in enumerate(cols, 1):
    hcell(p.cell(1, j), h)
row = 2
for w in snap['warehouses']:
    for r in w['rows']:
        base(p.cell(row, 1), w['name'])
        base(p.cell(row, 2), r['articleNum'], align='center')
        base(p.cell(row, 3), r['variant'])
        base(p.cell(row, 4), r['techSize'], align='center')
        base(p.cell(row, 5), r['nmID'], numfmt='0')
        base(p.cell(row, 6), r['stock'], align='right', numfmt='#,##0')
        base(p.cell(row, 7), r['perDay'], align='right', numfmt='0.00')
        base(p.cell(row, 8), ('∞' if r['daysToZero'] is None else r['daysToZero']), align='right', numfmt='0.0')
        base(p.cell(row, 9), r['reorderQty'], align='right', bold=True, numfmt='#,##0')
        base(p.cell(row, 10), r['status'])
        row += 1
p_last = row - 1
p.freeze_panes = 'A2'
p.auto_filter.ref = f'A1:J{p_last}'
for j, wdt in enumerate([18, 6, 22, 8, 12, 9, 9, 8, 9, 20], 1):
    p.column_dimensions[get_column_letter(j)].width = wdt
add_status_cf(p, 'J', 2, p_last)

# ── Лист «Пробный завоз» (по размерам) ──────────────────────────────────────
z = wb.create_sheet('Пробный завоз')
head = ['Арт', 'Цвет', 'Размер', 'nmID', 'Тип'] + whList + ['Итого']
for j, h in enumerate(head, 1):
    hcell(z.cell(1, j), h)
zrow = 2
zfirst = 6
for it in snap['seedGrid']:
    base(z.cell(zrow, 1), it['articleNum'], align='center')
    base(z.cell(zrow, 2), it['variant'])
    base(z.cell(zrow, 3), it['techSize'], align='center')
    base(z.cell(zrow, 4), it['nmID'], numfmt='0')
    base(z.cell(zrow, 5), it['kind'], align='center')
    sby = it['seedByWarehouse']
    for wi, wn in enumerate(whList):
        c = z.cell(zrow, zfirst + wi)
        if wn in sby:
            base(c, sby[wn], align='right', numfmt='#,##0')
        else:
            base(c, '✓', align='center', color='16875A')
    base(z.cell(zrow, zfirst + len(whList)), it['seedTotal'], align='right', bold=True, numfmt='#,##0')
    zrow += 1
z_last = zrow - 1
z.freeze_panes = 'D2'
z.auto_filter.ref = f'A1:{get_column_letter(zfirst + len(whList))}{z_last}'
for col, wdt in [('A', 6), ('B', 22), ('C', 9), ('D', 12), ('E', 10)]:
    z.column_dimensions[col].width = wdt
for wi in range(len(whList)):
    z.column_dimensions[get_column_letter(zfirst + wi)].width = 13
z.column_dimensions[get_column_letter(zfirst + len(whList))].width = 9
newF = PatternFill('solid', fgColor='E1F3E9'); refF = PatternFill('solid', fgColor='FAF0DA')
tr = f'E2:E{z_last}'
z.conditional_formatting.add(tr, CellIsRule(operator='equal', formula=['"новинка"'], fill=newF))
z.conditional_formatting.add(tr, CellIsRule(operator='equal', formula=['"докладка"'], fill=refF))

wb.save(OUT)
print('saved', os.path.relpath(OUT, os.getcwd()),
      '| подсорт', p_last - 1, '| сводная', sv_last - 1, '| завоз', z_last - 1)
