# scripts/fbs-cancels-xlsx.py — Excel по «Отказам по фулфилментам».
# Вход:  <REPORTS_OUTPUT_DIR>/fbs-cancels-service.json (снимок сервиса)
# Выход: <REPORTS_OUTPUT_DIR>/fbs-cancels.xlsx
# Листы: «Отказы по ФФ», «Причины потерь».
import os
import json
from openpyxl import Workbook
from xlsx_kit import autofit, style_header, autofilter

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_DIR = os.path.abspath(os.environ['REPORTS_OUTPUT_DIR']) if os.environ.get('REPORTS_OUTPUT_DIR') else os.path.join(REPO, 'reports-output')
SNAP = os.path.join(OUT_DIR, 'fbs-cancels-service.json')
OUT = os.path.join(OUT_DIR, 'fbs-cancels.xlsx')

with open(SNAP, encoding='utf-8') as f:
    s = json.load(f)

RED = 'B91C1C'  # заливка шапки листов «потерь»


def sheet(ws, head, rows, cells, rub_cols=(), pct_cols=()):
    ws.append(head)
    style_header(ws, fill=RED, wrap=True)
    for r in rows:
        ws.append(cells(r))
        rr = ws.max_row
        for j in rub_cols:
            ws.cell(rr, j).number_format = '#,##0 ₽'
        for j in pct_cols:
            ws.cell(rr, j).number_format = '0.0"%"'
    ws.freeze_panes = 'A2'
    autofilter(ws)
    autofit(ws)


wb = Workbook()
ws = wb.active
ws.title = 'Отказы по ФФ'
sheet(ws,
      ['Фулфилмент', 'Заданий', 'Выкуплено', 'Отказ ФФ', '% ФФ', 'Брак', 'Отказ клиента', 'Отмена 1-й час', 'В работе', 'Потери, ₽'],
      s.get('byFF', []),
      lambda r: [r.get('ff', ''), r.get('made', 0), r.get('sold', 0), r.get('sellerCancel', 0), r.get('sellerCancelPct', 0),
                 r.get('defect', 0), r.get('clientRefusal', 0), r.get('clientDecline', 0), r.get('inWork', 0), round(r.get('lostRub', 0))],
      rub_cols=(10,), pct_cols=(5,))

sheet(wb.create_sheet('Причины потерь'),
      ['Причина', 'Зона', 'Количество', 'Сумма, ₽'],
      s.get('reasons', []),
      lambda r: [r.get('ru', ''), {'ff': 'фулфилмент', 'client': 'клиент', 'other': 'прочее'}.get(r.get('blame'), r.get('blame', '')),
                 r.get('count', 0), round(r.get('rub', 0))],
      rub_cols=(4,))

wb.save(OUT)
print('OK', OUT)
