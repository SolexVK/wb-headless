#!/usr/bin/env python3
# Собирает Excel-отчёт из reports-output/burned-matrix-*.json:
# таблица NMID × склад (шт) + Итого шт + Себестоимость + Сумма ₽.
# Запуск: python3 scripts/build-matrix-xlsx.py [matrix.json] [out.xlsx]

import json, sys, glob, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

def latest():
    fs = sorted(glob.glob(os.path.join(ROOT, "reports-output", "burned-matrix-*.json")),
                key=os.path.getmtime, reverse=True)
    if not fs: sys.exit("нет burned-matrix-*.json — сначала node report-burned-matrix.mjs")
    return fs[0]

jpath = sys.argv[1] if len(sys.argv) > 1 else latest()
data = json.load(open(jpath, encoding="utf-8"))
warehouses = data["warehouses"]; rows = data["rows"]; cost = data["cost"]

out = sys.argv[2] if len(sys.argv) > 2 else os.path.join(
    ROOT, "reports-output", "burned-matrix-" +
    os.path.basename(jpath).replace("burned-matrix-", "").replace(".json", "") + ".xlsx")

# ── стили ──
PURPLE = "481173"; PURPLE2 = "7B2CBF"; FIRE = "E63946"; LIGHT = "EFE7F7"
hdr_fill = PatternFill("solid", fgColor=PURPLE)
wh_fill  = PatternFill("solid", fgColor=PURPLE2)
sum_fill = PatternFill("solid", fgColor="2A0A47")
tot_fill = PatternFill("solid", fgColor="D9C7EE")
white_b  = Font(bold=True, color="FFFFFF", size=10)
white_s  = Font(bold=True, color="FFFFFF", size=9)
bold     = Font(bold=True, size=10)
thin     = Side(style="thin", color="D0C0E8")
border   = Border(left=thin, right=thin, top=thin, bottom=thin)
center   = Alignment(horizontal="center", vertical="center", wrap_text=True)
right    = Alignment(horizontal="right")
left     = Alignment(horizontal="left")

wb = Workbook()
ws = wb.active
ws.title = "Потери по NMID"

# Заголовок-шапка отчёта
ncols = 4 + len(warehouses) + 3
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
c = ws.cell(1, 1, "Потери сгоревшего товара по артикулам (NMID) — склады Wildberries")
c.font = Font(bold=True, size=13, color=PURPLE); c.alignment = left
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
meta = "Снимок %s · себестоимость %d ₽/ед · основа: %s · всего %s ед / %s ₽" % (
    data["snapshotAt"][:10], cost, data["basis"],
    format(data["grandQty"], ",d").replace(",", " "),
    format(data["grandSumRub"], ",d").replace(",", " "))
c = ws.cell(2, 1, meta); c.font = Font(size=9, color="6A5A85"); c.alignment = left

HROW = 4
headers = ["NMID", "Артикул продавца", "Бренд", "Предмет"] + warehouses + ["Итого шт", "Себестоимость ₽", "Сумма ₽"]
for j, h in enumerate(headers, 1):
    cell = ws.cell(HROW, j, h)
    cell.font = white_s if 5 <= j <= 4 + len(warehouses) else white_b
    cell.fill = wh_fill if 5 <= j <= 4 + len(warehouses) else hdr_fill
    if j == ncols: cell.fill = PatternFill("solid", fgColor=FIRE)
    cell.alignment = center; cell.border = border
ws.row_dimensions[HROW].height = 64

r = HROW + 1
for row in rows:
    ws.cell(r, 1, row["nmId"]).alignment = right
    ws.cell(r, 2, row["vendorCode"]).alignment = left
    ws.cell(r, 3, row["brand"]).alignment = left
    ws.cell(r, 4, row["subjectName"]).alignment = left
    for k, wname in enumerate(warehouses):
        q = row["perWh"].get(wname, 0)
        cell = ws.cell(r, 5 + k, q if q else None)
        cell.alignment = right; cell.number_format = "#,##0"
    ws.cell(r, 5 + len(warehouses), row["totalQty"]).number_format = "#,##0"
    ws.cell(r, 5 + len(warehouses)).font = bold
    ws.cell(r, 6 + len(warehouses), cost).number_format = "#,##0"
    sc = ws.cell(r, 7 + len(warehouses), row["sumRub"])
    sc.number_format = "#,##0 ₽"; sc.font = bold
    for j in range(1, ncols + 1):
        ws.cell(r, j).border = border
    r += 1

# Итоговая строка
tot = ws.cell(r, 1, "ИТОГО"); tot.font = bold
for k, wname in enumerate(warehouses):
    cell = ws.cell(r, 5 + k, data["warehouseTotals"].get(wname, 0))
    cell.number_format = "#,##0"; cell.font = white_s; cell.fill = sum_fill; cell.alignment = right
ws.cell(r, 5 + len(warehouses), data["grandQty"]).number_format = "#,##0"
gsum = ws.cell(r, 7 + len(warehouses), data["grandSumRub"]); gsum.number_format = "#,##0 ₽"
for j in range(1, ncols + 1):
    cc = ws.cell(r, j); cc.border = border
    if j <= 4 or j >= 5 + len(warehouses): cc.fill = tot_fill; cc.font = bold
ws.row_dimensions[r].height = 20

# фильтры, закрепление, ширины
ws.auto_filter.ref = "A%d:%s%d" % (HROW, get_column_letter(ncols), r - 1)
ws.freeze_panes = "E%d" % (HROW + 1)
ws.column_dimensions["A"].width = 12
ws.column_dimensions["B"].width = 22
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 15
for k in range(len(warehouses)):
    ws.column_dimensions[get_column_letter(5 + k)].width = 9
ws.column_dimensions[get_column_letter(5 + len(warehouses))].width = 10
ws.column_dimensions[get_column_letter(6 + len(warehouses))].width = 13
ws.column_dimensions[get_column_letter(7 + len(warehouses))].width = 15

# ── лист «Итоги по складам» ──
ws2 = wb.create_sheet("Итоги по складам")
for j, h in enumerate(["Склад", "Сгорело, шт", "Сумма ₽"], 1):
    cell = ws2.cell(1, j, h); cell.font = white_b; cell.fill = hdr_fill; cell.alignment = center; cell.border = border
wt = sorted(data["warehouseTotals"].items(), key=lambda x: -x[1])
for i, (name, q) in enumerate(wt, 2):
    ws2.cell(i, 1, name).alignment = left
    ws2.cell(i, 2, q).number_format = "#,##0"
    ws2.cell(i, 3, q * cost).number_format = "#,##0 ₽"
    for j in range(1, 4): ws2.cell(i, j).border = border
tr = len(wt) + 2
ws2.cell(tr, 1, "ИТОГО").font = bold
ws2.cell(tr, 2, data["grandQty"]).number_format = "#,##0"; ws2.cell(tr, 2).font = bold
ws2.cell(tr, 3, data["grandSumRub"]).number_format = "#,##0 ₽"; ws2.cell(tr, 3).font = bold
for j in range(1, 4): ws2.cell(tr, j).fill = tot_fill; ws2.cell(tr, j).border = border
ws2.column_dimensions["A"].width = 32; ws2.column_dimensions["B"].width = 14; ws2.column_dimensions["C"].width = 16
ws2.freeze_panes = "A2"

wb.save(out)
print(out)
