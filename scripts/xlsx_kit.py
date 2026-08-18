# scripts/xlsx_kit.py — общие помощники сборки Excel-выгрузок FBS.
# Раньше autofit() копипастился 1:1 в четырёх *-xlsx.py, а оформление шапки
# (жирный + заливка + центрирование) — почти в каждом. Теперь одно место.
# Запускаются скрипты как `python3 scripts/<name>.py`, поэтому каталог scripts/
# уже на sys.path — импорт `from xlsx_kit import …` работает без правок пути.
from openpyxl.styles import Font, Alignment, PatternFill


def autofit(ws, cap=42, floor=9, pad=2):
    """Ширина колонок по содержимому: max(floor, длина+pad), но не шире cap."""
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(cap, max(floor, width + pad))


def style_header(ws, row=1, fill=None, wrap=False):
    """Оформить строку-шапку: жирный (+ белый на заливке), центрирование (+ перенос)."""
    font = Font(bold=True, color='FFFFFF') if fill else Font(bold=True)
    pf = PatternFill('solid', fgColor=fill) if fill else None
    align = Alignment(horizontal='center', vertical='center', wrap_text=True) if wrap \
        else Alignment(horizontal='center')
    for c in ws[row]:
        c.font = font
        if pf:
            c.fill = pf
        c.alignment = align


def autofilter(ws):
    """Автофильтр по всей таблице (если есть хотя бы одна строка данных)."""
    if ws.max_row > 1:
        ws.auto_filter.ref = f'A1:{ws.cell(1, ws.max_column).column_letter}{ws.max_row}'
