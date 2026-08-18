# scripts/fbs-logistics-xlsx.py — Excel по «Логистике» (сроки сборки и доставки).
# Вход:  <REPORTS_OUTPUT_DIR>/fbs-logistics-service.json (снимок сервиса)
# Выход: <REPORTS_OUTPUT_DIR>/fbs-logistics.xlsx
# Листы: «Сборка по ФФ», «Долгие сборки», «Доставка по ФФ»,
#        «Доставка по регионам», «Доставка по дням». Время — в часах.
import os
import json
from openpyxl import Workbook
from xlsx_kit import autofit, style_header, autofilter

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_DIR = os.path.abspath(os.environ['REPORTS_OUTPUT_DIR']) if os.environ.get('REPORTS_OUTPUT_DIR') else os.path.join(REPO, 'reports-output')
SNAP = os.path.join(OUT_DIR, 'fbs-logistics-service.json')
OUT = os.path.join(OUT_DIR, 'fbs-logistics.xlsx')

with open(SNAP, encoding='utf-8') as f:
    s = json.load(f)

INDIGO = '4338CA'  # цвет заливки шапки листов логистики


def sheet(ws, head, rows, cells, hour_cols=()):
    ws.append(head)
    style_header(ws, fill=INDIGO, wrap=True)
    for r in rows:
        ws.append(cells(r))
        rr = ws.max_row
        for j in hour_cols:
            ws.cell(rr, j).number_format = '0.0'
    ws.freeze_panes = 'A2'
    autofilter(ws)
    autofit(ws)


A = s.get('assembly', {}) or {}
D = s.get('delivery', {}) or {}

wb = Workbook()
ws = wb.active
ws.title = 'Сборка по ФФ'
sheet(ws, ['ФФ отгрузки', 'Сделано', 'Обработано', 'В работе', 'Среднее, ч', 'Медиана, ч', 'p90, ч', 'Макс, ч', 'Критич.'],
      A.get('byFF', []),
      lambda r: [r.get('ff', ''), r.get('made', 0), r.get('processed', 0), r.get('pending', 0),
                 r.get('avgHours', 0), r.get('medianHours', 0), r.get('p90Hours', 0), r.get('maxHours', 0), r.get('criticalCount', 0)],
      hour_cols=(5, 6, 7, 8))

sheet(wb.create_sheet('Долгие сборки'), ['ФФ', 'Артикул', 'Время сборки, ч', 'Создан', 'Отгружен'],
      A.get('critical', []),
      lambda r: [r.get('ff', ''), r.get('article', ''), r.get('hours', 0),
                 str(r.get('createdAt', ''))[:16].replace('T', ' '), str(r.get('closedAt', ''))[:16].replace('T', ' ')],
      hour_cols=(3,))

sheet(wb.create_sheet('Доставка по ФФ'), ['ФФ отгрузки', 'Выкупов', 'Среднее, ч', 'Медиана, ч', 'p90, ч', 'Мин, ч', 'Макс, ч'],
      D.get('byFF', []),
      lambda r: [r.get('ff', ''), r.get('count', 0), r.get('avgHours', 0), r.get('medianHours', 0), r.get('p90Hours', 0), r.get('minHours', 0), r.get('maxHours', 0)],
      hour_cols=(3, 4, 5, 6, 7))

sheet(wb.create_sheet('Доставка по регионам'), ['Округ', 'Регион покупателя', 'Выкупов', 'Среднее, ч', 'Медиана, ч'],
      D.get('byRegion', []),
      lambda r: [r.get('okrug', ''), r.get('region', ''), r.get('count', 0), r.get('avgHours', 0), r.get('medianHours', 0)],
      hour_cols=(4, 5))

sheet(wb.create_sheet('Доставка по дням'), ['Дата отгрузки', 'Выкупов', 'Среднее, ч'],
      D.get('byDay', []),
      lambda r: [r.get('date', ''), r.get('count', 0), r.get('avgHours', 0)],
      hour_cols=(3,))

wb.save(OUT)
print('OK', OUT)
