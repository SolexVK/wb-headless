# -*- coding: utf-8 -*-
"""
Генератор Excel-книги из JSON сводного отчёта по остаткам WB.

Вход  — reports-output/wb-stock-<дата>.json (см. report-wb-stock.js).
Выход — .xlsx с листами: Сводка, FBO по складам, FBS по складам,
        «В пути и возвраты».

Запуск:
    pip install openpyxl
    python build_report.py <report.json> [output.xlsx]

Зависимости: openpyxl (pandas не нужен — данные уже агрегированы в JSON).
"""
import sys
import json
import argparse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

FONT = "Arial"
NAVY, BLUE, LBLUE, LGREY = "1F3864", "2E75B6", "D6E4F0", "F2F4F7"
GREEN, RED, AMBER, WHITE = "2E7D32", "C0392B", "B7791F", "FFFFFF"
INT = "#,##0"

_thin = Side(style="thin", color="D0D5DD")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def cell(ws, ref, value=None, *, bold=False, size=11, color="000000",
         fill=None, align="left", valign="center", num=None, wrap=False,
         border=False, italic=False):
    c = ws[ref]
    if value is not None:
        c.value = value
    c.font = Font(name=FONT, bold=bold, size=size, color=color, italic=italic)
    c.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap)
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    if num:
        c.number_format = num
    if border:
        c.border = BORDER
    return c


def kpi_card(ws, col_letter, row, label, value, clr):
    c2 = chr(ord(col_letter) + 1)
    ws.merge_cells(f"{col_letter}{row}:{c2}{row}")
    cell(ws, f"{col_letter}{row}", label, bold=True, size=9, color="667085",
         fill=LGREY, align="center", border=True)
    ws.merge_cells(f"{col_letter}{row+1}:{c2}{row+1}")
    cell(ws, f"{col_letter}{row+1}", value, bold=True, size=18, color=clr,
         align="center", border=True, num=INT)


def sheet_header(ws, last_col, title, subtitle=None):
    ws.sheet_view.showGridLines = False
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=last_col)
    cell(ws, "B1", title, bold=True, size=15, color=WHITE, fill=NAVY)
    ws.row_dimensions[1].height = 26
    if subtitle:
        ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=last_col)
        cell(ws, "B2", subtitle, italic=True, size=9, color="667085")


def table(ws, start_row, columns, rows, totals=None):
    """columns: list of (key, title, width, is_num). Возвращает (hr, first, last)."""
    hr = start_row
    for j, (key, title, width, is_num) in enumerate(columns, 2):
        L = get_column_letter(j)
        ws.column_dimensions[L].width = width
        cell(ws, f"{L}{hr}", title, bold=True, color=WHITE, fill=BLUE,
             align="center" if is_num else "left", border=True, size=9)
    first = hr + 1
    for idx, r in enumerate(rows):
        rr = first + idx
        fillc = LGREY if idx % 2 else WHITE
        for j, (key, title, width, is_num) in enumerate(columns, 2):
            L = get_column_letter(j)
            cell(ws, f"{L}{rr}", r.get(key, 0 if is_num else ""),
                 num=INT if is_num else None, align="center" if is_num else "left",
                 border=True, size=10, fill=fillc)
    last = first + len(rows) - 1
    if totals is not None:
        tr = last + 1
        cell(ws, f"B{tr}", "ИТОГО", bold=True, color=WHITE, fill=NAVY, border=True)
        for j, (key, title, width, is_num) in enumerate(columns, 2):
            if j == 2:
                continue
            L = get_column_letter(j)
            # nmID — идентификатор, не суммируем.
            if is_num and key != "nmId" and last >= first:
                cell(ws, f"{L}{tr}", f"=SUM({L}{first}:{L}{last})", num=INT,
                     align="center", border=True, bold=True, color=WHITE, fill=NAVY, size=10)
            else:
                cell(ws, f"{L}{tr}", "", border=True, fill=NAVY)
    return hr, first, last


def build(report, out_path):
    wb = openpyxl.Workbook()
    gen = report.get("generatedAt", "")
    t = report.get("totals", {})
    warnings = report.get("warnings", [])

    # ─────────────── Сводка ───────────────
    ws = wb.active
    ws.title = "Сводка"
    sub = f"На {gen[:19].replace('T', ' ')} · артикулов: {t.get('articles', 0)}"
    if warnings:
        sub += " · ⚠ есть предупреждения (см. JSON): " + "; ".join(warnings)[:120]
    sheet_header(ws, 9, "ОСТАТКИ WB — СВОДКА ПО АРТИКУЛУ ПРОДАВЦА", sub)

    kpi_card(ws, "B", 4, "FBO доступно", t.get("fboAvailable", 0), BLUE)
    kpi_card(ws, "D", 4, "FBS остаток", t.get("fbs", 0), BLUE)
    kpi_card(ws, "F", 4, "Итого доступно", t.get("totalAvailable", 0), NAVY)
    kpi_card(ws, "H", 4, "В пути к клиенту", t.get("inWayToClient", 0), GREEN)
    kpi_card(ws, "B", 7, "В пути от клиента (возвраты)", t.get("inWayFromClient", 0), AMBER)
    kpi_card(ws, "D", 7, "FBO всего на складах", t.get("fboFull", 0), "667085")

    cols = [
        ("seller", "Артикул продавца", 34, False),
        ("nmId", "nmID", 12, True),
        ("fboAvailable", "FBO дост.", 11, True),
        ("fboFull", "FBO всего", 11, True),
        ("fbs", "FBS", 10, True),
        ("totalAvailable", "Итого дост.", 12, True),
        ("inWayToClient", "→ клиенту", 11, True),
        ("inWayFromClient", "← возврат", 11, True),
    ]
    ws.column_dimensions["A"].width = 2
    hr, first, last = table(ws, 10, cols, report.get("byArticle", []), totals=t)
    ws.freeze_panes = f"C{first}"

    # график: топ-15 по «Итого доступно» (FBO+FBS)
    n = min(15, len(report.get("byArticle", [])))
    if n:
        ch = BarChart(); ch.type = "bar"; ch.title = "Топ-15 артикулов по доступному остатку"
        ch.height, ch.width = max(8, 0.55 * n), 16; ch.legend = None
        total_col = 2 + 5  # 'totalAvailable' — 6-я колонка данных → столбец G(7)
        ch.add_data(Reference(ws, min_col=total_col, min_row=first, max_row=first + n - 1))
        ch.set_categories(Reference(ws, min_col=2, min_row=first, max_row=first + n - 1))
        ch.series[0].graphicalProperties.solidFill = BLUE
        ws.add_chart(ch, "K4")

    # ─────────────── FBO по складам ───────────────
    wf = wb.create_sheet("FBO по складам")
    sheet_header(wf, 6, "FBO — ОСТАТКИ ПО СКЛАДАМ WB",
                 "Склады Wildberries. «В пути» — логистика FBO.")
    wf.column_dimensions["A"].width = 2
    fcols = [
        ("warehouse", "Склад", 34, False),
        ("available", "Доступно", 12, True),
        ("full", "Всего", 12, True),
        ("inWayToClient", "→ клиенту", 12, True),
        ("inWayFromClient", "← возврат", 12, True),
    ]
    _, ff, fl = table(wf, 4, fcols, report.get("fboByWarehouse", []), totals=True)
    wf.freeze_panes = f"C{ff}"
    m = len(report.get("fboByWarehouse", []))
    if m:
        ch = BarChart(); ch.type = "bar"; ch.title = "Доступно по складам FBO"
        ch.height, ch.width = max(8, 0.5 * m), 14; ch.legend = None
        ch.add_data(Reference(wf, min_col=3, min_row=ff, max_row=ff + m - 1))
        ch.set_categories(Reference(wf, min_col=2, min_row=ff, max_row=ff + m - 1))
        ch.series[0].graphicalProperties.solidFill = BLUE
        wf.add_chart(ch, "H4")

    # ─────────────── FBS по складам ───────────────
    wbs = wb.create_sheet("FBS по складам")
    sheet_header(wbs, 4, "FBS — ОСТАТКИ ПО ФУЛФИЛМЕНТ-СКЛАДАМ ПРОДАВЦА",
                 "Собственные склады продавца (Маркетплейс API).")
    wbs.column_dimensions["A"].width = 2
    scols = [
        ("warehouse", "Склад", 34, False),
        ("amount", "Остаток, шт", 14, True),
    ]
    fbs_rows = report.get("fbsByWarehouse", [])
    table(wbs, 4, scols, fbs_rows, totals=True)
    if not fbs_rows:
        cell(wbs, "B7", "Нет данных FBS (склады не заданы или токен без категории «Маркетплейс»).",
             italic=True, color="667085")

    # ─────────────── В пути и возвраты ───────────────
    wt = wb.create_sheet("В пути и возвраты")
    sheet_header(wt, 4, "В ПУТИ К КЛИЕНТУ И ВОЗВРАТЫ",
                 "Артикулы, где есть движение: «в пути к клиенту» или «в пути от клиента» (возвраты).")
    wt.column_dimensions["A"].width = 2
    moving = [r for r in report.get("byArticle", [])
              if r.get("inWayToClient", 0) or r.get("inWayFromClient", 0)]
    moving.sort(key=lambda r: (-r.get("inWayFromClient", 0), -r.get("inWayToClient", 0)))
    tcols = [
        ("seller", "Артикул продавца", 40, False),
        ("inWayToClient", "→ клиенту", 14, True),
        ("inWayFromClient", "← возврат", 14, True),
    ]
    table(wt, 4, tcols, moving, totals={"inWayToClient": t.get("inWayToClient", 0),
                                        "inWayFromClient": t.get("inWayFromClient", 0)})

    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.save(out_path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("input", help="JSON-отчёт wb-stock-<дата>.json")
    ap.add_argument("output", nargs="?", default="wb-stock-report.xlsx")
    args = ap.parse_args()
    with open(args.input, encoding="utf-8") as f:
        report = json.load(f)
    build(report, args.output)
    print("saved", args.output)


if __name__ == "__main__":
    main()
