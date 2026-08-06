# -*- coding: utf-8 -*-
"""
Генератор интерактивного Excel-дашборда динамики остатков.

Вход: .xlsx с листом, где строки — комбинации (Артикул, Размер, Склад),
а столбцы 4..N — остатки по дням (заголовки-даты). Формат совпадает с
выгрузкой «Остатки по дням».

Запуск:
    python build_dashboard.py <input.xlsx> [output.xlsx] [--sheet "Остатки по дням"]

Результат — рабочая книга с листами:
    Дашборд      — 3 фильтра (Артикул/Размер/Склад, «ВСЕ» = без фильтра),
                   KPI и график динамики на живых SUMIFS (пересчёт при смене фильтра);
    По артикулам — таблица по всем артикулам × дням + изм./%, цветовая шкала, топ-15;
    По размерам  — сводка по размерам + график начало/конец;
    По складам   — сводка по складам + топ-15;
    Данные       — исходный снимок (на него ссылаются формулы дашборда).

Зависимости: openpyxl, pandas.
"""
import sys
import argparse
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.datavalidation import DataValidation


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("input")
    ap.add_argument("output", nargs="?", default="dashboard.xlsx")
    ap.add_argument("--sheet", default=None, help="Имя листа с данными (по умолчанию — первый)")
    args = ap.parse_args()

    df = pd.read_excel(args.input, sheet_name=args.sheet or 0)
    build(df, args.output)
    print("saved", args.output)


# ------------------------------------------------------------------ styles
FONT = "Arial"
NAVY, BLUE, LBLUE, LGREY = "1F3864", "2E75B6", "D6E4F0", "F2F4F7"
GREEN, RED, WHITE = "2E7D32", "C0392B", "FFFFFF"
INT, PCT, SGN = "#,##0", "0.0%;-0.0%;0.0%", "+#,##0;-#,##0;0"

_thin = Side(style="thin", color="D0D5DD")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def cell(ws, ref, value=None, *, bold=False, size=11, color="000000",
         fill=None, align="left", valign="center", num=None, wrap=False,
         border=False, italic=False):
    c = ws[ref]
    if value is not None:
        c.value = value
    c.font = Font(name=FONT, bold=bold, size=size, color=color, italic=italic)
    c.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap)
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    if num:
        c.number_format = num
    if border:
        c.border = BORDER
    return c


def build(df, out_path):
    cols = list(df.columns)
    key_cols = cols[:3]              # Артикул, Размер, Склад
    dcols = cols[3:]                 # даты
    ART, SIZE, WH = key_cols
    NDAY = len(dcols)
    first, last = dcols[0], dcols[-1]
    NROW = len(df)
    DATA_LAST = NROW + 1

    # orderings (values live via formulas / precomputed; only order comes from here)
    art = df.groupby(ART)[dcols].sum(); art["chg"] = art[last] - art[first]
    articles_by_change = art.sort_values("chg").index.tolist()
    sz = df.groupby(SIZE)[dcols].sum(); sz["chg"] = sz[last] - sz[first]
    sizes_by_stock = sz.sort_values(first, ascending=False).index.tolist()
    wh = df.groupby(WH)[dcols].sum(); wh["chg"] = wh[last] - wh[first]
    wh_by_change = wh.sort_values("chg").index.tolist()
    all_articles = sorted(df[ART].unique().tolist())
    all_sizes = sorted(df[SIZE].unique().tolist())
    all_wh = sorted(df[WH].unique().tolist())

    wb = openpyxl.Workbook()

    # ---------------- Данные (raw) ----------------
    wsd = wb.active
    wsd.title = "Данные"
    for j, col in enumerate(cols, 1):
        cell(wsd, f"{get_column_letter(j)}1", str(col), bold=True, color=WHITE,
             fill=NAVY, align="center", border=True)
    for i, row in enumerate(df.itertuples(index=False), 2):
        for j, val in enumerate(row, 1):
            c = wsd.cell(row=i, column=j, value=val)
            c.font = Font(name=FONT, size=10)
            if j >= 4:
                c.number_format = INT
                c.alignment = Alignment(horizontal="center")
    wsd.freeze_panes = "D2"
    wsd.column_dimensions["A"].width = 34
    wsd.column_dimensions["B"].width = 10
    wsd.column_dimensions["C"].width = 30
    for j in range(4, 4 + NDAY):
        wsd.column_dimensions[get_column_letter(j)].width = 12

    DN = "Данные"
    A_RNG = f"{DN}!$A$2:$A${DATA_LAST}"
    B_RNG = f"{DN}!$B$2:$B${DATA_LAST}"
    C_RNG = f"{DN}!$C$2:$C${DATA_LAST}"

    def day_rng(i):
        L = get_column_letter(4 + i)
        return f"{DN}!${L}$2:${L}${DATA_LAST}"

    # ---------------- Списки (hidden, dropdowns) ----------------
    wsl = wb.create_sheet("Списки")
    wsl["A1"], wsl["C1"], wsl["E1"] = "Артикулы", "Размеры", "Склады"
    wsl["A2"] = wsl["C2"] = wsl["E2"] = "ВСЕ"
    for i, a in enumerate(all_articles, 3):
        wsl[f"A{i}"] = a
    for i, s in enumerate(all_sizes, 3):
        wsl[f"C{i}"] = s
    for i, w in enumerate(all_wh, 3):
        wsl[f"E{i}"] = w
    ART_LAST, SZ_LAST, WH_LAST = 2 + len(all_articles), 2 + len(all_sizes), 2 + len(all_wh)
    wsl.sheet_state = "hidden"

    # ---------------- Дашборд ----------------
    ws = wb.create_sheet("Дашборд")
    ws.sheet_view.showGridLines = False
    for col, w in {"A": 2, "B": 20, "C": 16, "D": 14, "E": 14, "F": 14,
                   "G": 14, "H": 14, "I": 14, "J": 14, "K": 2}.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("B2:J2")
    cell(ws, "B2", "ДИНАМИКА ОСТАТКОВ  ·  интерактивный дашборд",
         bold=True, size=18, color=WHITE, fill=NAVY)
    ws.row_dimensions[2].height = 30
    ws.merge_cells("B3:J3")
    cell(ws, "B3", f"Период {first} – {last} · {df[ART].nunique()} артикулов · "
                   f"{df[SIZE].nunique()} размеров · {df[WH].nunique()} складов",
         size=10, color=WHITE, fill=BLUE)
    ws.row_dimensions[3].height = 18

    cell(ws, "B5", "ФИЛЬТРЫ", bold=True, color=NAVY)
    filters = [("B6", "Артикул", "C6", f"Списки!$A$2:$A${ART_LAST}"),
               ("B7", "Размер", "C7", f"Списки!$C$2:$C${SZ_LAST}"),
               ("B8", "Склад", "C8", f"Списки!$E$2:$E${WH_LAST}")]
    for lref, lab, vref, listrng in filters:
        cell(ws, lref, lab, bold=True, color="344054", fill=LGREY, align="right", border=True)
        cell(ws, vref, "ВСЕ", bold=True, color=NAVY, fill=LBLUE, align="left", border=True)
        dv = DataValidation(type="list", formula1=listrng, allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(ws[vref])
    ws.merge_cells("D6:J6")
    cell(ws, "D6", "◄ Выберите значения из списков. «ВСЕ» = без фильтра. "
                   "График и KPI ниже пересчитываются автоматически.",
         italic=True, size=10, color="667085", wrap=True)
    ws.merge_cells("D7:J8")

    # criteria helpers ("<>" matches any non-empty cell — no wildcard settings needed)
    ws["M6"] = '=IF(C6="ВСЕ","<>",C6)'
    ws["M7"] = '=IF(C7="ВСЕ","<>",C7)'
    ws["M8"] = '=IF(C8="ВСЕ","<>",C8)'
    for r in (6, 7, 8):
        ws[f"M{r}"].font = Font(name=FONT, size=10, color=WHITE)
    ws.column_dimensions["M"].hidden = True

    # series table feeding the chart
    cell(ws, "B22", "Динамика по дням (по выбранному фильтру)", bold=True, color=NAVY)
    cell(ws, "B23", "Дата", bold=True, color=WHITE, fill=BLUE, align="center", border=True)
    cell(ws, "C23", "Остаток, шт", bold=True, color=WHITE, fill=BLUE, align="center", border=True)
    SER_FIRST = 24
    for i, d in enumerate(dcols):
        r = SER_FIRST + i
        cell(ws, f"B{r}", str(d), align="center", border=True, size=10)
        f = (f"=SUMIFS({day_rng(i)},{A_RNG},$M$6,{B_RNG},$M$7,{C_RNG},$M$8)")
        cell(ws, f"C{r}", f, align="center", border=True, num=INT)
    SER_LAST = SER_FIRST + NDAY - 1

    kpis = [("B10", "Остаток на начало", f"=C{SER_FIRST}", INT, BLUE),
            ("D10", "Остаток на конец", f"=C{SER_LAST}", INT, BLUE),
            ("F10", "Изменение, шт", f"=C{SER_LAST}-C{SER_FIRST}", SGN, NAVY),
            ("H10", "Изменение, %", f"=IFERROR((C{SER_LAST}-C{SER_FIRST})/C{SER_FIRST},0)", PCT, NAVY)]
    spans = {"B10": "B10:C10", "D10": "D10:E10", "F10": "F10:G10", "H10": "H10:I10"}
    vspans = {"B11": "B11:C11", "D11": "D11:E11", "F11": "F11:G11", "H11": "H11:I11"}
    for (ref, lab, formula, fmt, clr), vkey in zip(kpis, vspans):
        ws.merge_cells(spans[ref])
        cell(ws, ref, lab, bold=True, size=10, color="667085", fill=LGREY, align="center", border=True)
        ws.merge_cells(vspans[vkey])
        cell(ws, vkey, formula, bold=True, size=20, color=clr, fill=WHITE,
             align="center", border=True, num=fmt)
    ws.row_dimensions[10].height = 18
    ws.row_dimensions[11].height = 30

    mini = [("B13", "Минимум", f"=MIN(C{SER_FIRST}:C{SER_LAST})"),
            ("D13", "Максимум", f"=MAX(C{SER_FIRST}:C{SER_LAST})"),
            ("F13", "Среднее", f"=ROUND(AVERAGE(C{SER_FIRST}:C{SER_LAST}),0)"),
            ("H13", "Строк в выборке",
             f"=SUMPRODUCT((COUNTIFS({A_RNG},$M$6,{B_RNG},$M$7,{C_RNG},$M$8)))")]
    for ref, lab, formula in mini:
        c2 = chr(ord(ref[0]) + 1) + "13"
        ws.merge_cells(f"{ref}:{c2}")
        cell(ws, ref, lab, bold=True, size=9, color="667085", align="center", border=True, fill=LGREY)
        vr, vr2 = ref[0] + "14", chr(ord(ref[0]) + 1) + "14"
        ws.merge_cells(f"{vr}:{vr2}")
        cell(ws, vr, formula, bold=True, size=13, color="344054", align="center", border=True, num=INT)

    chart = LineChart()
    chart.title = "Остаток по дням"
    chart.style = 2
    chart.height, chart.width = 8.2, 20
    chart.y_axis.title = "шт"
    chart.y_axis.majorGridlines = ChartLines()
    chart.x_axis.delete = chart.y_axis.delete = False
    data = Reference(ws, min_col=3, min_row=23, max_row=SER_LAST)
    cats = Reference(ws, min_col=2, min_row=24, max_row=SER_LAST)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = BLUE
    s.graphicalProperties.line.width = 28000
    ws.add_chart(chart, "D16")
    ws.freeze_panes = "A4"

    # ---------------- summary sheets (precomputed values from the snapshot) ----------------
    def build_table(title, keys, agg, key_header, key_width, chart_title, chart_n):
        wt = wb.create_sheet(title)
        wt.sheet_view.showGridLines = False
        wt.column_dimensions["A"].width = 2
        wt.column_dimensions["B"].width = key_width
        for j in range(3, 3 + NDAY):
            wt.column_dimensions[get_column_letter(j)].width = 11
        chg_col, pct_col = 3 + NDAY, 4 + NDAY
        wt.column_dimensions[get_column_letter(chg_col)].width = 12
        wt.column_dimensions[get_column_letter(pct_col)].width = 11

        wt.merge_cells(start_row=1, start_column=2, end_row=1, end_column=pct_col)
        cell(wt, "B1", title.upper(), bold=True, size=15, color=WHITE, fill=NAVY)
        wt.row_dimensions[1].height = 26
        wt.merge_cells(start_row=2, start_column=2, end_row=2, end_column=pct_col)
        cell(wt, "B2", "Значения рассчитаны из листа «Данные». "
                       "Отсортировано по величине изменения за период.",
             italic=True, size=9, color="667085")

        hr = 3
        cell(wt, f"B{hr}", key_header, bold=True, color=WHITE, fill=BLUE, border=True)
        for i, d in enumerate(dcols):
            cell(wt, f"{get_column_letter(3+i)}{hr}", str(d), bold=True, color=WHITE,
                 fill=BLUE, align="center", border=True, size=9)
        cell(wt, f"{get_column_letter(chg_col)}{hr}", "Изм., шт", bold=True, color=WHITE,
             fill=BLUE, align="center", border=True, size=9)
        cell(wt, f"{get_column_letter(pct_col)}{hr}", "Изм., %", bold=True, color=WHITE,
             fill=BLUE, align="center", border=True, size=9)

        start = hr + 1
        for idx, key in enumerate(keys):
            r = start + idx
            fillc = LGREY if idx % 2 else WHITE
            cell(wt, f"B{r}", key, size=10, border=True, fill=fillc)
            vals = [int(agg.loc[key, d]) for d in dcols]
            for i in range(NDAY):
                cell(wt, f"{get_column_letter(3+i)}{r}", vals[i], num=INT,
                     align="center", border=True, size=10, fill=fillc)
            change = vals[-1] - vals[0]
            pct = (change / vals[0]) if vals[0] else 0
            cell(wt, f"{get_column_letter(chg_col)}{r}", change, num=SGN,
                 align="center", border=True, bold=True, size=10, fill=fillc)
            cell(wt, f"{get_column_letter(pct_col)}{r}", pct, num=PCT,
                 align="center", border=True, size=10, fill=fillc)
        end = start + len(keys) - 1

        tr = end + 1
        cL, pL = get_column_letter(chg_col), get_column_letter(pct_col)
        fL, lL = get_column_letter(3), get_column_letter(3 + NDAY - 1)
        cell(wt, f"B{tr}", "ИТОГО", bold=True, color=WHITE, fill=NAVY, border=True)
        for i in range(NDAY):
            L = get_column_letter(3 + i)
            cell(wt, f"{L}{tr}", f"=SUM({L}{start}:{L}{end})", num=INT, align="center",
                 border=True, bold=True, color=WHITE, fill=NAVY, size=10)
        cell(wt, f"{cL}{tr}", f"={lL}{tr}-{fL}{tr}", num=SGN, align="center",
             border=True, bold=True, color=WHITE, fill=NAVY, size=10)
        cell(wt, f"{pL}{tr}", f"=IFERROR(({lL}{tr}-{fL}{tr})/{fL}{tr},0)", num=PCT,
             align="center", border=True, bold=True, color=WHITE, fill=NAVY, size=10)

        wt.conditional_formatting.add(
            f"{cL}{start}:{cL}{end}",
            ColorScaleRule(start_type="min", start_color="F8696B",
                           mid_type="num", mid_value=0, mid_color="FFFFFF",
                           end_type="max", end_color="63BE7B"))
        wt.freeze_panes = f"C{start}"

        ch = BarChart()
        ch.type = "bar"
        ch.title = chart_title
        ch.height, ch.width = max(8, 0.55 * chart_n), 16
        ch.legend = None
        rr0, rr1 = start, start + chart_n - 1
        ch.add_data(Reference(wt, min_col=chg_col, min_row=rr0, max_row=rr1), titles_from_data=False)
        ch.set_categories(Reference(wt, min_col=2, min_row=rr0, max_row=rr1))
        ch.series[0].graphicalProperties.solidFill = RED
        wt.add_chart(ch, f"{get_column_letter(pct_col + 2)}{hr}")
        return wt, start, end, chg_col, pct_col

    build_table("По артикулам", articles_by_change, art, ART, 34,
                "Топ-15 артикулов по снижению остатка", min(15, len(articles_by_change)))
    build_table("По складам", wh_by_change, wh, WH, 30,
                "Топ-15 складов по снижению остатка", min(15, len(wh_by_change)))
    wsz, s0, s1, _, szpct = build_table("По размерам", sizes_by_stock, sz, SIZE, 12,
                                        "Изменение остатка по размерам, шт", len(sizes_by_stock))
    cc = BarChart(); cc.type = "col"
    cc.title = "Остаток на начало и конец периода по размерам"
    cc.height, cc.width = 9, 20
    cc.y_axis.title = "шт"
    cc.add_data(Reference(wsz, min_col=3, min_row=s0 - 1, max_row=s1), titles_from_data=True)
    cc.add_data(Reference(wsz, min_col=3 + NDAY - 1, min_row=s0 - 1, max_row=s1), titles_from_data=True)
    cc.set_categories(Reference(wsz, min_col=2, min_row=s0, max_row=s1))
    cc.series[0].graphicalProperties.solidFill = BLUE
    cc.series[1].graphicalProperties.solidFill = NAVY
    wsz.add_chart(cc, f"{get_column_letter(szpct + 2)}{s0 + 18}")

    order = ["Дашборд", "По артикулам", "По размерам", "По складам", "Данные", "Списки"]
    wb._sheets.sort(key=lambda s: order.index(s.title))
    wb.active = 0
    # recalc every formula on open (headless LibreOffice recalc not required)
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.save(out_path)


if __name__ == "__main__":
    main()
